import requests
import time
import os
import sys
import random
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# 设置控制台编码为UTF-8
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

class BilibiliVideoScraper:
    def __init__(self):
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Referer': 'https://www.bilibili.com',
            'Accept': 'application/json',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8'
        }
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.covers_dir = os.path.join(self.base_dir, 'covers')
        os.makedirs(self.covers_dir, exist_ok=True)

    def search_videos(self, keyword='鸣潮', page_size=50):
        """搜索B站视频，支持重试机制"""
        search_url = "https://api.bilibili.com/x/web-interface/search/type"
        params = {
            'search_type': 'video',
            'keyword': keyword,
            'page': 1,
            'page_size': page_size
        }

        max_retries = 3
        for retry in range(max_retries):
            try:
                print(f"正在搜索 '{keyword}' 相关视频... (尝试 {retry + 1}/{max_retries})")

                if retry > 0:
                    delay = random.uniform(2, 5)
                    print(f"等待 {delay:.1f} 秒后重试...")
                    time.sleep(delay)

                response = requests.get(search_url, headers=self.headers, params=params, timeout=15)

                if response.status_code == 412:
                    print(f"请求被阻止(412)，等待后重试...")
                    time.sleep(3)
                    continue

                if response.status_code == 200:
                    data = response.json()
                    if data.get('code') == 0:
                        results = data.get('data', {}).get('result', [])
                        print(f"成功获取到 {len(results)} 个视频")
                        return results
                    elif data.get('code') == -509:
                        print(f"请求过于频繁，请稍后再试")
                        time.sleep(5)
                        continue
                    else:
                        print(f"API返回错误: {data.get('message')}")
                else:
                    print(f"请求失败，状态码: {response.status_code}")

            except Exception as e:
                print(f"搜索出错: {e}")
                time.sleep(2)

        print("搜索失败，已达到最大重试次数")
        return []

    def parse_video_info(self, search_results):
        """解析搜索结果中的视频信息"""
        videos = []

        for item in search_results:
            try:
                bvid = item.get('bvid', '')
                if not bvid:
                    continue

                title = item.get('title', '')
                title = title.replace('<em class="keyword">', '').replace('</em>', '')

                play_str = item.get('play', '0')
                play_count = self.parse_number(play_str)

                video = {
                    'title': title,
                    'author': item.get('author', ''),
                    'bvid': bvid,
                    'play_count': play_count,
                    'play_str': play_str,
                    'duration': item.get('duration', ''),
                    'pubdate': item.get('pubdate', 0),
                    'video_review': item.get('video_review', 0),
                    'likes': item.get('like', 0),
                    'pic': item.get('pic', ''),
                    'url': f"https://www.bilibili.com/video/{bvid}"
                }
                videos.append(video)

            except Exception as e:
                print(f"解析视频信息出错: {e}")
                continue

        return videos

    def parse_number(self, num_str):
        """解析数字字符串（处理万、千等单位）"""
        if isinstance(num_str, (int, float)):
            return int(num_str)

        try:
            num_str = str(num_str).replace(',', '').strip()
            if '万' in num_str:
                return int(float(num_str.replace('万', '')) * 10000)
            elif '千' in num_str:
                return int(float(num_str.replace('千', '')) * 1000)
            else:
                return int(num_str)
        except (ValueError, TypeError):
            return 0

    def download_cover(self, pic_url, bvid):
        """下载视频封面"""
        filename = f"{bvid}.jpg"
        filepath = os.path.join(self.covers_dir, filename)

        if os.path.exists(filepath):
            print(f"  封面已存在: {filename}")
            return filepath

        urls_to_try = []
        if pic_url:
            if pic_url.startswith('//'):
                pic_url = 'https:' + pic_url
            if pic_url.startswith('http'):
                urls_to_try.append(pic_url)
                urls_to_try.append(pic_url.replace('http://', 'https://'))

        for url in urls_to_try:
            try:
                response = requests.get(url, headers=self.headers, timeout=10)
                if response.status_code == 200 and len(response.content) > 1000:
                    with open(filepath, 'wb') as f:
                        f.write(response.content)
                    print(f"  下载封面成功: {filename}")
                    return filepath
            except Exception:
                continue

        print(f"  封面下载失败: {bvid}")
        return None

    def format_date(self, timestamp):
        """格式化时间戳"""
        try:
            return datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            return ''

    def create_excel(self, videos):
        """创建Excel表格，videos已按播放量降序排列"""
        if not videos:
            print("没有视频数据可导出")
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "鸣潮热门视频TOP10"

        headers = ['排名', '视频标题', '作者', 'BV号', '播放量', '时长', '发布时间', '弹幕数', '点赞数', '视频链接']

        # 写入表头
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 设置列宽
        col_widths = [6, 40, 15, 12, 10, 8, 18, 8, 8, 45]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # 写入数据
        for row_idx, video in enumerate(videos, start=2):
            rank = row_idx - 1
            ws.cell(row=row_idx, column=1, value=rank)
            ws.cell(row=row_idx, column=2, value=video['title'])
            ws.cell(row=row_idx, column=3, value=video['author'])
            ws.cell(row=row_idx, column=4, value=video['bvid'])
            ws.cell(row=row_idx, column=5, value=video['play_str'])
            ws.cell(row=row_idx, column=6, value=video['duration'])
            ws.cell(row=row_idx, column=7, value=self.format_date(video['pubdate']))
            ws.cell(row=row_idx, column=8, value=video['video_review'])
            ws.cell(row=row_idx, column=9, value=video['likes'])
            ws.cell(row=row_idx, column=10, value=video['url'])

            link_cell = ws.cell(row=row_idx, column=10)
            link_cell.hyperlink = video['url']
            link_cell.style = "Hyperlink"

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"鸣潮热门视频TOP10_{timestamp}.xlsx"
        output_path = os.path.join(self.base_dir, filename)
        wb.save(output_path)
        print(f"\n[OK] Excel表格已生成: {filename}")

        return output_path

    def get_video_info_by_bvid(self, bvid):
        """通过BV号获取视频详情"""
        url = "https://api.bilibili.com/x/web-interface/view"
        params = {'bvid': bvid}

        try:
            response = requests.get(url, headers=self.headers, params=params, timeout=15)
            if response.status_code != 200:
                print(f"请求失败，状态码: {response.status_code}")
                return None

            data = response.json()
            if data.get('code') != 0:
                print(f"API返回错误: {data.get('message', '未知错误')}")
                return None

            info = data['data']
            return {
                'title': info.get('title', ''),
                'author': info.get('owner', {}).get('name', ''),
                'bvid': bvid,
                'pic': info.get('pic', ''),
                'url': f"https://www.bilibili.com/video/{bvid}"
            }
        except Exception as e:
            print(f"获取视频信息出错: {e}")
            return None

    def run_bv_lookup(self):
        """BV号查询模式：输入BV号，下载对应封面"""
        print("=" * 60)
        print("请输入BV号（例如：BV1xx411c7mD）")
        print("=" * 60)
        print()

        while True:
            bvid = input("BV号: ").strip()
            if not bvid:
                print("BV号不能为空，请重新输入")
                continue
            if not bvid.upper().startswith('BV'):
                print("BV号格式不正确，应以BV开头，请重新输入")
                continue
            break

        print(f"\n正在查询 {bvid} 的视频信息...")
        info = self.get_video_info_by_bvid(bvid)

        if not info:
            print("\n⚠ 查询失败，请检查BV号是否正确")
            return

        print(f"\n视频标题: {info['title']}")
        print(f"作者: {info['author']}")
        print(f"链接: {info['url']}")

        print(f"\n正在下载封面...")
        filepath = self.download_cover(info['pic'], bvid)

        if filepath:
            print(f"\n✓ 封面已保存至: {filepath}")
        else:
            print("\n⚠ 封面下载失败")

    def run(self):
        """执行完整流程"""
        print("=" * 60)
        print("B站鸣潮热门视频采集工具")
        print("=" * 60)
        print()

        # 1. 搜索视频
        search_results = self.search_videos()
        if not search_results:
            print("\n⚠ 搜索失败或无结果")
            print("\n提示：")
            print("1. 检查网络连接")
            print("2. B站可能限制了API访问，请稍后再试")
            return

        # 2. 解析视频信息
        videos = self.parse_video_info(search_results)
        if not videos:
            print("\n⚠ 未解析到视频信息")
            return

        print(f"\n共解析到 {len(videos)} 个视频")

        # 3. 按播放量排序，取TOP 10
        sorted_videos = sorted(videos, key=lambda x: x['play_count'], reverse=True)[:10]

        print("\n" + "=" * 60)
        print("TOP 10 热门视频")
        print("=" * 60)

        for rank, video in enumerate(sorted_videos, start=1):
            print(f"\n{rank}. {video['title'][:30]}...")
            print(f"   作者: {video['author']}")
            print(f"   播放量: {video['play_str']}")
            print(f"   BV号: {video['bvid']}")

        # 4. 下载封面图片
        print("\n" + "=" * 60)
        print("开始下载封面图片...")
        print("=" * 60)

        for video in sorted_videos:
            self.download_cover(video['pic'], video['bvid'])
            time.sleep(0.5)

        # 5. 生成Excel表格
        print("\n" + "=" * 60)
        print("生成Excel表格...")
        print("=" * 60)

        excel_path = self.create_excel(sorted_videos)
        if excel_path:
            self.generate_summary(sorted_videos, os.path.basename(excel_path))

        print("\n" + "=" * 60)
        print("✓ 采集完成！")
        print("=" * 60)
        print(f"\n生成的文件：")
        print(f"1. Excel表格: {os.path.basename(excel_path) if excel_path else '无'}")
        print(f"2. 封面图片: covers/ 目录")

    def generate_summary(self, videos, excel_filename):
        """生成汇总报告"""
        summary_filename = excel_filename.replace('.xlsx', '_报告.txt')
        summary_path = os.path.join(self.base_dir, summary_filename)

        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write("鸣潮热门视频TOP10 采集报告\n")
            f.write(f"采集时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 60 + "\n\n")

            for rank, video in enumerate(videos, start=1):
                f.write(f"【排名 {rank}】\n")
                f.write(f"标题: {video['title']}\n")
                f.write(f"作者: {video['author']}\n")
                f.write(f"BV号: {video['bvid']}\n")
                f.write(f"播放量: {video['play_str']}\n")
                f.write(f"时长: {video['duration']}\n")
                f.write(f"弹幕数: {video['video_review']}\n")
                f.write(f"点赞数: {video['likes']}\n")
                f.write(f"链接: {video['url']}\n")
                f.write("-" * 60 + "\n\n")

        print(f"✓ 汇总报告已生成: {summary_filename}")


def main():
    scanf = input
    print("=" * 60)
    print("B站视频工具")
    print("=" * 60)
    print("1. 搜索鸣潮热门视频 TOP10（默认）")
    print("2. 输入 BV 号获取视频封面")
    print("-" * 60)

    choice = scanf("请选择功能 (1/2): ").strip()
    print()

    scraper = BilibiliVideoScraper()
    if choice == '2':
        scraper.run_bv_lookup()
    else:
        scraper.run()

if __name__ == "__main__":
    main()
